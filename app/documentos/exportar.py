"""
Export utilities for Lista Mestra: Excel (openpyxl), PDF (reportlab), CSV.
Each function receives a list of Documento objects and returns a BytesIO/str.
"""

import csv
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from app.utils.datetime_utils import agora_brasilia

_HEADERS = [
    'Código', 'Título', 'Tipo', 'Revisão', 'Data Aprovação',
    'Status', 'Dist. Técnica', 'Dist. Admin.',
]

_EXT_HEADERS = [
    'Identificação', 'Título', 'Órgão Emissor / Categoria',
    'Revisão', 'Distribuição', 'Situação',
]

_HEADER_HEX = 'FF1A2035'
_HEADER_TEXT = 'FFFFFFFF'
_ALT_ROW_HEX = 'FFF0F4FF'


def _doc_to_row(doc) -> list:
    return [
        doc.codigo,
        doc.titulo,
        doc.tipo_documento,
        doc.revisao_formatada,
        doc.data_aprovacao.strftime('%d/%m/%Y') if doc.data_aprovacao else '—',
        doc.status,
        'Sim' if doc.distribuicao_tecnica else 'Não',
        'Sim' if doc.distribuicao_administrativa else 'Não',
    ]


def _ext_doc_to_row(ext) -> list:
    """Convert a DocumentoExterno to the _EXT_HEADERS column list."""
    revisao = ext.revisao.strip() if ext.revisao and ext.revisao.strip() else 'N/A'
    if ext.distribuicao_tecnica and ext.distribuicao_administrativa:
        dist = 'Técnica / Administrativa'
    elif ext.distribuicao_tecnica:
        dist = 'Técnica'
    elif ext.distribuicao_administrativa:
        dist = 'Administrativa'
    else:
        dist = '—'
    return [
        ext.codigo or '—',
        ext.titulo,
        ext.orgao_emissor or '—',
        revisao,
        dist,
        ext.status,
    ]


# ── Excel ──────────────────────────────────────────────────────────────────────

def gerar_excel_lista_mestra(documentos, externos=None, consultas=None, versao_sw=None) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lista Mestra'

    thin = Side(style='thin', color='FFD0D8E4')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title row (merged A1:J1)
    ws.merge_cells('A1:J1')
    tc = ws['A1']
    tc.value = (
        'LISTA MESTRA DE DOCUMENTOS – CSV Cascavel\n'
        f'Gerado em: {agora_brasilia().strftime("%d/%m/%Y %H:%M")} Horário de Brasília   |   '
        f'Total vigentes: {len(documentos)}'
    )
    tc.font = Font(bold=True, size=11, color=_HEADER_TEXT)
    tc.fill = PatternFill('solid', fgColor=_HEADER_HEX)
    tc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 42

    # ── Header row
    hdr_font = Font(bold=True, color=_HEADER_TEXT, size=9)
    hdr_fill = PatternFill('solid', fgColor='FF2A3550')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, text in enumerate(_HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=text)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
    ws.row_dimensions[2].height = 28

    # ── Data rows
    alt_fill = PatternFill('solid', fgColor=_ALT_ROW_HEX)
    data_align = Alignment(vertical='center', wrap_text=True)

    for row_i, doc in enumerate(documentos, 3):
        fill = alt_fill if row_i % 2 == 0 else None
        for col_i, value in enumerate(_doc_to_row(doc), 1):
            cell = ws.cell(row=row_i, column=col_i, value=value)
            cell.border = border
            cell.alignment = data_align
            cell.font = Font(size=9)
            if fill:
                cell.fill = fill

    # ── Column widths
    col_widths = [14, 42, 18, 9, 14, 14, 13, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A3'

    # ── External documents (appended on a separate sheet) ───────────────────
    if externos:
        ws_ext = wb.create_sheet(title='Doc. Externos')
        ext_hdr_font = Font(bold=True, color=_HEADER_TEXT, size=9)
        ext_hdr_fill = PatternFill('solid', fgColor='FF1E6B7B')
        ext_hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_ext = Side(style='thin', color='FFD0D8E4')
        border_ext = Border(
            left=thin_ext, right=thin_ext, top=thin_ext, bottom=thin_ext
        )

        for col, text in enumerate(_EXT_HEADERS, 1):
            cell = ws_ext.cell(row=1, column=col, value=text)
            cell.font = ext_hdr_font
            cell.fill = ext_hdr_fill
            cell.alignment = ext_hdr_align
            cell.border = border_ext
        ws_ext.row_dimensions[1].height = 24

        for row_i, ext in enumerate(externos, 2):
            fill = PatternFill('solid', fgColor=_ALT_ROW_HEX) if row_i % 2 == 0 else None
            for col_i, value in enumerate(_ext_doc_to_row(ext), 1):
                cell = ws_ext.cell(row=row_i, column=col_i, value=value)
                cell.border = border_ext
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.font = Font(size=9)
                if fill:
                    cell.fill = fill

        ext_col_widths = [20, 50, 24, 10, 26, 12]
        for i, w in enumerate(ext_col_widths, 1):
            ws_ext.column_dimensions[get_column_letter(i)].width = w

    # ── Consultas Remotas (separate sheet) ────────────────────────────────────
    if consultas:
        _MESES_NOME = [
            (1,'JAN'),(2,'FEV'),(3,'MAR'),(4,'ABR'),
            (5,'MAI'),(6,'JUN'),(7,'JUL'),(8,'AGO'),
            (9,'SET'),(10,'OUT'),(11,'NOV'),(12,'DEZ'),
        ]
        ws_cr = wb.create_sheet(title='Consultas Remotas')
        cr_hdr_font = Font(bold=True, color=_HEADER_TEXT, size=9)
        cr_hdr_fill = PatternFill('solid', fgColor='FF1A2035')
        cr_hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cr_border = Border(
            left=Side(style='thin', color='FFD0D8E4'),
            right=Side(style='thin', color='FFD0D8E4'),
            top=Side(style='thin', color='FFD0D8E4'),
            bottom=Side(style='thin', color='FFD0D8E4'),
        )
        cr_headers = [
            'Mês', 'Status 1ª Quinzena', 'Data 1ª Quinzena', 'Responsável 1ª',
            'Status 2ª Quinzena', 'Data 2ª Quinzena', 'Responsável 2ª',
        ]
        for col, text in enumerate(cr_headers, 1):
            cell = ws_cr.cell(row=1, column=col, value=text)
            cell.font = cr_hdr_font
            cell.fill = cr_hdr_fill
            cell.alignment = cr_hdr_align
            cell.border = cr_border
        ws_cr.row_dimensions[1].height = 24

        mapa = {(c.mes, c.quinzena): c for c in consultas}
        for row_i, (num_mes, nome_mes) in enumerate(_MESES_NOME, 2):
            fill = PatternFill('solid', fgColor=_ALT_ROW_HEX) if row_i % 2 == 0 else None
            q1 = mapa.get((num_mes, 1))
            q2 = mapa.get((num_mes, 2))
            row_data = [
                nome_mes,
                'Verificado' if q1 and q1.verificado else 'Pendente',
                q1.verificado_em.strftime('%d/%m/%Y') if q1 and q1.verificado_em else '',
                q1.verificado_por.nome if q1 and q1.verificado_por else '',
                'Verificado' if q2 and q2.verificado else 'Pendente',
                q2.verificado_em.strftime('%d/%m/%Y') if q2 and q2.verificado_em else '',
                q2.verificado_por.nome if q2 and q2.verificado_por else '',
            ]
            for col_i, value in enumerate(row_data, 1):
                cell = ws_cr.cell(row=row_i, column=col_i, value=value)
                cell.border = cr_border
                cell.alignment = Alignment(vertical='center')
                cell.font = Font(size=9)
                if fill:
                    cell.fill = fill

        for i, w in enumerate([8, 18, 16, 20, 18, 16, 20], 1):
            ws_cr.column_dimensions[get_column_letter(i)].width = w

    # ── Versão de Software (separate sheet) ──────────────────────────────────
    if versao_sw:
        ws_vs = wb.create_sheet(title='Versão Software')
        vs_hdr_font = Font(bold=True, color=_HEADER_TEXT, size=9)
        vs_hdr_fill = PatternFill('solid', fgColor='FF1A2035')
        vs_hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        vs_border = Border(
            left=Side(style='thin', color='FFD0D8E4'),
            right=Side(style='thin', color='FFD0D8E4'),
            top=Side(style='thin', color='FFD0D8E4'),
            bottom=Side(style='thin', color='FFD0D8E4'),
        )
        vs_headers = ['Equipamento', 'Software', 'Versão']
        for col, text in enumerate(vs_headers, 1):
            cell = ws_vs.cell(row=1, column=col, value=text)
            cell.font = vs_hdr_font
            cell.fill = vs_hdr_fill
            cell.alignment = vs_hdr_align
            cell.border = vs_border
        ws_vs.row_dimensions[1].height = 24

        for row_i, r in enumerate(versao_sw, 2):
            fill = PatternFill('solid', fgColor=_ALT_ROW_HEX) if row_i % 2 == 0 else None
            row_data = [r.equipamento, r.software, r.versao]
            for col_i, value in enumerate(row_data, 1):
                cell = ws_vs.cell(row=row_i, column=col_i, value=value)
                cell.border = vs_border
                cell.alignment = Alignment(vertical='center')
                cell.font = Font(size=9)
                if fill:
                    cell.fill = fill

        for i, w in enumerate([35, 35, 15], 1):
            ws_vs.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── PDF ────────────────────────────────────────────────────────────────────────

def gerar_pdf_lista_mestra(documentos, cfg=None, externos=None, consultas=None, versao_sw=None) -> io.BytesIO:
    """Generate a PDF for the Lista Mestra.

    If *cfg* (a ListaMestraConfig instance) is supplied the PDF includes a
    document-style header with logo, code, revision and signatories.
    If *externos* is supplied, active external documents are appended.
    """
    import os
    try:
        from flask import current_app
        _logo_path = os.path.join(
            current_app.root_path, 'static', 'images', 'logo.png'
        )
    except RuntimeError:
        _logo_path = None

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1 * cm, leftMargin=1 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        title='Lista Mestra – CSV Cascavel',
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'SGQTitle', parent=styles['Title'],
        fontSize=13, spaceAfter=4,
        textColor=colors.HexColor('#1a2035'),
    )
    sub_style = ParagraphStyle(
        'SGQSub', parent=styles['Normal'],
        fontSize=8, textColor=colors.HexColor('#555555'), spaceAfter=8,
    )
    cell_style = ParagraphStyle(
        'SGQCell', parent=styles['Normal'],
        fontSize=7, leading=8.5,
        spaceBefore=0, spaceAfter=0,
    )

    elements = []

    # ── Document header (when cfg is provided) ─────────────────────────────────
    if cfg is not None:
        try:
            from reportlab.platypus import Image as RLImage
            _has_image = _logo_path and os.path.exists(_logo_path)
        except ImportError:
            _has_image = False

        logo_cell = (
            RLImage(_logo_path, width=3.5 * cm, height=1.5 * cm, kind='proportional')
            if _has_image else Paragraph('<b>CSV Cascavel</b>', cell_style)
        )

        rev_str = f'Rev{cfg.revisao_num:02d}'
        data_aprov = (
            cfg.data_aprovacao.strftime('%d/%m/%Y') if cfg.data_aprovacao else '—'
        )
        elab = cfg.elaborado_por.nome if cfg.elaborado_por else '—'
        rev_por = cfg.revisado_por.nome if cfg.revisado_por else '—'
        aprov = cfg.aprovado_por.nome if cfg.aprovado_por else '—'

        hdr_style = TableStyle([
            ('BOX', (0, 0), (-1, -1), 0.6, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.4, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('SPAN', (0, 0), (0, 2)),    # logo spans 3 rows
            ('SPAN', (1, 0), (2, 0)),    # title spans 2 cols
            ('ALIGN', (1, 0), (2, 0), 'CENTER'),
            ('FONTNAME', (1, 0), (2, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (1, 0), (2, 0), 11),
        ])

        hdr_table = Table(
            [
                [logo_cell, Paragraph(f'<b>{cfg.titulo}</b>', cell_style), ''],
                [''        , Paragraph(f'<b>Código:</b> {cfg.codigo}', cell_style),
                             Paragraph(f'<b>Revisão:</b> {rev_str}', cell_style)],
                [''        , Paragraph(f'<b>Data de aprovação:</b> {data_aprov}', cell_style),
                             Paragraph(f'<b>Gerado em:</b> {agora_brasilia().strftime("%d/%m/%Y")}', cell_style)],
            ],
            colWidths=[3.5 * cm, None, None],
            hAlign='LEFT',
        )
        hdr_table.setStyle(hdr_style)
        elements.append(hdr_table)

        sig_table = Table(
            [[
                Paragraph(f'<b>Elaborado por:</b> {elab}', cell_style),
                Paragraph(f'<b>Revisado por:</b> {rev_por}', cell_style),
                Paragraph(f'<b>Aprovado por:</b> {aprov}', cell_style),
            ]],
            hAlign='LEFT',
        )
        sig_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 0.6, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.4, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(sig_table)
        elements.append(Spacer(1, 0.2 * cm))
    else:
        title_text = 'LISTA MESTRA DE DOCUMENTOS – CSV Cascavel'
        elements += [
            Paragraph(title_text, title_style),
            Paragraph(
                f'Gerado em: {agora_brasilia().strftime("%d/%m/%Y %H:%M")} Horário de Brasília  |  '
                f'Total de documentos vigentes: {len(documentos)}',
                sub_style,
            ),
            Spacer(1, 0.3 * cm),
        ]

    # ── Table data (using Paragraph for auto word wrapping) ─────────────────
    data = [[Paragraph(h, cell_style) for h in _HEADERS]]
    for d in documentos:
        row = _doc_to_row(d)
        data.append([Paragraph(str(v), cell_style) for v in row])

    # Wider column widths for better readability on landscape A4 (usable ~27.7cm)
    col_w = [2.2, 8.0, 3.2, 1.5, 2.4, 2.4, 2.0, 2.0]
    col_w_pt = [w * cm for w in col_w]

    table = Table(data, colWidths=col_w_pt, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a2035')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7.5),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # Data
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
            colors.white, colors.HexColor('#f0f4ff')
        ]),
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#c0c8d8')),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))

    # Set row heights to auto by wrapping text — ReportLab will size them
    elements.append(table)

    # ── External documents section ─────────────────────────────────────────
    if externos:
        ext_title_style = ParagraphStyle(
            'SGQExtTitle', parent=styles['Normal'],
            fontSize=9, textColor=colors.HexColor('#1E6B7B'),
            spaceBefore=14, spaceAfter=4, fontName='Helvetica-Bold',
        )
        elements.append(
            Paragraph('DOCUMENTOS EXTERNOS – VIGENTES', ext_title_style)
        )

        ext_data = [[Paragraph(h, cell_style) for h in _EXT_HEADERS]]
        for ext in externos:
            ext_data.append([Paragraph(str(v), cell_style) for v in _ext_doc_to_row(ext)])

        ext_col_w = [3.5, 8.0, 4.0, 1.6, 3.5, 2.0]
        ext_col_w_pt = [w * cm for w in ext_col_w]

        ext_table = Table(ext_data, colWidths=ext_col_w_pt, repeatRows=1)
        ext_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E6B7B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7.5),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                colors.white, colors.HexColor('#e8f7f9')
            ]),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#c0c8d8')),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(ext_table)

    # ── Consultas Remotas section in PDF ──────────────────────────────────────
    if consultas:
        _MESES_NOME = [
            (1,'JAN'),(2,'FEV'),(3,'MAR'),(4,'ABR'),
            (5,'MAI'),(6,'JUN'),(7,'JUL'),(8,'AGO'),
            (9,'SET'),(10,'OUT'),(11,'NOV'),(12,'DEZ'),
        ]
        elements.append(Spacer(1, 0.35 * cm))
        from reportlab.platypus import Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        styles = getSampleStyleSheet()
        cr_title_style = styles['Heading2']
        elements.append(Paragraph('Controle de Consultas Remotas', cr_title_style))

        mapa = {(c.mes, c.quinzena): c for c in consultas}
        cr_data = [[
            'Mês',
            '1ª Quinzena', 'Data 1ª', 'Responsável 1ª',
            '2ª Quinzena', 'Data 2ª', 'Responsável 2ª',
        ]]
        for num_mes, nome_mes in _MESES_NOME:
            q1 = mapa.get((num_mes, 1))
            q2 = mapa.get((num_mes, 2))
            cr_data.append([
                nome_mes,
                'Verificado' if q1 and q1.verificado else 'Pendente',
                q1.verificado_em.strftime('%d/%m/%Y') if q1 and q1.verificado_em else '—',
                q1.verificado_por.nome if q1 and q1.verificado_por else '—',
                'Verificado' if q2 and q2.verificado else 'Pendente',
                q2.verificado_em.strftime('%d/%m/%Y') if q2 and q2.verificado_em else '—',
                q2.verificado_por.nome if q2 and q2.verificado_por else '—',
            ])

        cr_col_w = [1.5, 2.2, 2.0, 3.2, 2.2, 2.0, 3.2]
        cr_col_w_pt = [w * cm for w in cr_col_w]
        cr_table = Table(cr_data, colWidths=cr_col_w_pt, repeatRows=1)
        cr_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A2035')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7.5),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                colors.white, colors.HexColor('#d1fae5')
            ]),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#c0c8d8')),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(cr_table)

    # ── Versão de Software section in PDF ────────────────────────────────────
    if versao_sw:
        elements.append(Spacer(1, 0.35 * cm))
        from reportlab.platypus import Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        styles = getSampleStyleSheet()
        vs_title_style = styles['Heading2']
        elements.append(Paragraph('Controle de Versão de Software', vs_title_style))

        vs_data = [['Equipamento', 'Software', 'Versão']]
        for r in versao_sw:
            vs_data.append([r.equipamento, r.software, r.versao])

        vs_col_w = [6.0, 6.0, 3.0]
        vs_col_w_pt = [w * cm for w in vs_col_w]
        vs_table = Table(vs_data, colWidths=vs_col_w_pt, repeatRows=1)
        vs_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A2035')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7.5),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                colors.white, colors.HexColor('#d1fae5')
            ]),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#c0c8d8')),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(vs_table)

    # Build PDF with a canvas that draws running headers with page X/Y
    try:
        from app.utils.html_pdf import NumberedCanvasFactory
        meta_for_header = {
            'codigo': getattr(cfg, 'codigo', None) if cfg is not None else None,
            'titulo': getattr(cfg, 'titulo', 'Lista Mestra de Documentos') if cfg is not None else 'Lista Mestra de Documentos',
            'revisao': getattr(cfg, 'revisao_num', None) if cfg is not None else None,
        }
        canvas_maker = NumberedCanvasFactory(meta_for_header) or None
    except Exception:
        canvas_maker = None

    if canvas_maker:
        doc.build(elements, canvasmaker=canvas_maker)
    else:
        doc.build(elements)

    buffer.seek(0)
    return buffer


# ── CSV ────────────────────────────────────────────────────────────────────────

def gerar_csv_lista_mestra(documentos, externos=None, consultas=None, versao_sw=None) -> str:
    """Return a CSV string with UTF-8 BOM (compatible with Excel)."""
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow(_HEADERS)
    for doc in documentos:
        writer.writerow(_doc_to_row(doc))
    # Append external documents section
    if externos:
        writer.writerow([])
        writer.writerow(['--- DOCUMENTOS EXTERNOS ---'])
        writer.writerow(_EXT_HEADERS)
        for ext in externos:
            writer.writerow(_ext_doc_to_row(ext))
    # Append consultas remotas section
    if consultas:
        _MESES_NOME = [
            (1,'JAN'),(2,'FEV'),(3,'MAR'),(4,'ABR'),
            (5,'MAI'),(6,'JUN'),(7,'JUL'),(8,'AGO'),
            (9,'SET'),(10,'OUT'),(11,'NOV'),(12,'DEZ'),
        ]
        writer.writerow([])
        writer.writerow(['--- CONSULTAS REMOTAS ---'])
        writer.writerow([
            'Mês', 'Status 1ª Quinzena', 'Data 1ª', 'Responsável 1ª',
            'Status 2ª Quinzena', 'Data 2ª', 'Responsável 2ª',
        ])
        mapa = {(c.mes, c.quinzena): c for c in consultas}
        for num_mes, nome_mes in _MESES_NOME:
            q1 = mapa.get((num_mes, 1))
            q2 = mapa.get((num_mes, 2))
            writer.writerow([
                nome_mes,
                'Verificado' if q1 and q1.verificado else 'Pendente',
                q1.verificado_em.strftime('%d/%m/%Y') if q1 and q1.verificado_em else '',
                q1.verificado_por.nome if q1 and q1.verificado_por else '',
                'Verificado' if q2 and q2.verificado else 'Pendente',
                q2.verificado_em.strftime('%d/%m/%Y') if q2 and q2.verificado_em else '',
                q2.verificado_por.nome if q2 and q2.verificado_por else '',
            ])
    # Append versão de software section
    if versao_sw:
        writer.writerow([])
        writer.writerow(['--- CONTROLE DE VERSÃO DE SOFTWARE ---'])
        writer.writerow(['Equipamento', 'Software', 'Versão'])
        for r in versao_sw:
            writer.writerow([r.equipamento, r.software, r.versao])
    # Prepend BOM for Excel compatibility
    return '\ufeff' + output.getvalue()
